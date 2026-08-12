"""Normalise the filled workbook: drop dead rows/blocks, rebuild sheet layout.

    python -m scripts.normalize_xlsx --xlsx results/PACLock_baseline_matrix_filled.xlsx

Runs after fill_xlsx.py, and edits the filled copy in place -- the input
workbook is never touched.

Content removals
----------------
1. **Retired model rows** (LaBraM from-scratch). Its runs and configs are gone,
   so the row could only stay blank.

2. **Empty external-reference blocks.** Five corpora have no comparable
   published numbers -- Sleep-EDF, ISRUC, PhysioNet-MI and BCI-IV-2a because no
   baseline paper reports them, TUSZ and FACED because the ones that exist use a
   different task construction. Their block held one sentence saying so; the
   sentence is preserved as a note under the results table, where it explains
   rather than dangles under an empty heading. TUAB, TUEV and CHB-MIT keep their
   blocks (15, 17 and 14 real anchor rows).

3. **All-empty metric columns.** Some sheets carry a trailing "Δ… vs PACLock"
   column that was never populated.

Layout rebuild
--------------
The workbook's borders were already uneven -- some model rows boxed, some with
only a bottom edge, some none -- and deleting rows made it worse by leaving
part-bordered blanks behind. Rather than patch individual cells, the results
table is re-drawn from scratch: every cell from the header row to the last model
row gets the same thin grid, one fixed row height, and one alignment rule
(labels left, numbers centred). Row heights are reset rather than preserved
because the inherited ones (13.75 to 123.8) clipped two-line model names.
"""

from __future__ import annotations

import argparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATASET_SHEETS = ["TUAB", "TUEV", "TUSZ", "CHB-MIT", "Sleep-EDF", "ISRUC",
                  "PhysioNet-MI", "FACED", "BCI-IV-2a"]

# Nothing is retired at the moment. Kept as a hook: a model dropped from the
# matrix leaves a blank row behind unless its label is listed here.
DROP_ROW_LABELS: set[str] = set()

ANCHOR_HEADING = "外部参考"
FROZEN_HEADING = "已核查预处理协议"

THIN = Side(style="thin", color="B0B0B0")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()

HEADER_FILL = PatternFill("solid", fgColor="E8F0E4")
GROUP_FILL = PatternFill("solid", fgColor="F5F5F5")
TITLE_FILL = PatternFill("solid", fgColor="E8F0E4")

TITLE_SIZE = 15.0
SUBTITLE_SIZE = 10.0
SUBTITLE_H = 22.0

ROW_H = 20.0
HEADER_H = 28.0
TITLE_H = 26.0

COL_WIDTHS = [24, 30, 12, 17, 17, 17, 17]


def find_row(ws, needle, col=1, start=1):
    for r in range(start, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and needle in v:
            return r
    return None


# Column A headings that open a block below the model table. Detection must key
# on these rather than on "column B went blank": BCI-IV-2a's protocol block
# follows the table with no separating row AND fills column B with parameter
# values, so a blank-column-B scan ran straight through it to the sheet end.
BLOCK_HEADINGS = (ANCHOR_HEADING, FROZEN_HEADING, "冻结。")


def results_span(ws):
    """(header_row, last_model_row) of the upper results table."""
    hdr = find_row(ws, "模型", col=2)
    if hdr is None:
        return None, None
    last = hdr
    for r in range(hdr + 1, ws.max_row + 1):
        group = ws.cell(row=r, column=1).value
        if isinstance(group, str) and any(h in group for h in BLOCK_HEADINGS):
            break
        label = ws.cell(row=r, column=2).value
        if isinstance(label, str) and label.strip():
            last = r
        elif r - last > 3:
            break
    return hdr, last


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    n_rows = n_blocks = n_cols = 0

    for name in DATASET_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]

        # ---- 1. retired model rows ------------------------------------- #
        while True:
            hit = next((r for r in range(1, ws.max_row + 1)
                        if str(ws.cell(row=r, column=2).value or "").strip()
                        in DROP_ROW_LABELS), None)
            if hit is None:
                break
            ws.delete_rows(hit)
            n_rows += 1

        # ---- 2. everything below the results table --------------------- #
        # The sheets carried two blocks under the model table: the published
        # "external reference" anchors and a transcription of the frozen
        # protocol. Both are dropped -- the anchors live in
        # configs/published_reference.json and the protocol in docs/PROTOCOLS.md,
        # which are the versioned sources; duplicating them in the workbook only
        # created a second copy to keep in sync. What remains is one table per
        # corpus and nothing else.
        _hdr, _last = results_span(ws)
        if _hdr is not None and ws.max_row > _last:
            # Drop the ranges from the collection directly. ws.unmerge_cells()
            # walks the member cells and raises KeyError when a merge covers
            # columns the sheet never materialised, which several of these
            # full-width banner merges do.
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row > _last:
                    ws.merged_cells.ranges.remove(rng)
            n_blocks += ws.max_row - _last
            ws.delete_rows(_last + 1, ws.max_row - _last)

        hdr, last = results_span(ws)
        if hdr is None:
            continue

        # ---- 3. blank rows left inside the results table ---------------- #
        r = hdr + 1
        while r <= last:
            row_empty = all(
                not str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, 8))
            if row_empty:
                ws.delete_rows(r)
                last -= 1
                n_rows += 1
            else:
                r += 1

        # ---- 4. all-empty metric columns -------------------------------- #
        for col in range(7, 3, -1):
            head = ws.cell(row=hdr, column=col).value
            if not head:
                continue
            body = [ws.cell(row=r, column=col).value for r in range(hdr + 1, last + 1)]
            if not any(str(v or "").strip() for v in body):
                ws.delete_cols(col)
                n_cols += 1

        # ---- 5. redraw the results table -------------------------------- #
        ncol = max(c for c in range(1, 8)
                   if ws.cell(row=hdr, column=c).value) if hdr else 6
        for r in range(hdr, last + 1):
            ws.row_dimensions[r].height = HEADER_H if r == hdr else ROW_H
            for c in range(1, ncol + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = GRID
                if r == hdr:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=True)
                elif c == 1:
                    cell.font = Font(bold=True, size=10)
                    cell.fill = GROUP_FILL
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c == 2:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ---- 6. title block, widths, panes ------------------------------ #
        # The FACED sheet is generated by add_faced_sheet.py rather than
        # inherited from the input workbook, so it never picked up the title
        # styling the other eight carry: no A1:G1 / A2:G2 merge, no fill, no
        # font size, and a half-width colon in the subtitle. Restyle the title
        # block on every sheet so the source of a sheet stops being visible.
        ncol_title = ncol
        for row, size, height, fill in (
            (1, TITLE_SIZE, TITLE_H, TITLE_FILL),
            (2, SUBTITLE_SIZE, SUBTITLE_H, None),
        ):
            span = f"A{row}:{get_column_letter(ncol_title)}{row}"
            if not any(str(rng) == span for rng in ws.merged_cells.ranges):
                for rng in list(ws.merged_cells.ranges):
                    if rng.min_row == row:
                        ws.merged_cells.ranges.remove(rng)
                ws.merge_cells(span)
            cell = ws.cell(row=row, column=1)
            if isinstance(cell.value, str):
                cell.value = cell.value.replace(":", "：")
            cell.font = Font(bold=(row == 1), size=size)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if fill is not None:
                cell.fill = fill
            ws.row_dimensions[row].height = height
        # Width only the columns the table actually uses. Setting all seven
        # left a wide blank G beside every table, and the sheets disagreed on H
        # (13.0 on the generated FACED sheet, ~8.7 on the inherited ones), so
        # anything past the last real column is reset to the default instead.
        for i in range(1, ncol + 1):
            ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]
        for i in range(ncol + 1, 12):
            letter = get_column_letter(i)
            if letter in ws.column_dimensions:
                del ws.column_dimensions[letter]

        # Blank rows immediately above and below the table inherited partial
        # borders from the original workbook (a left+top edge with no bottom,
        # which renders as a stray line). They are spacers, so clear them
        # outright rather than trying to complete the box.
        for r in list(range(max(1, hdr - 2), hdr)) + list(range(last + 1, last + 3)):
            if r < 1 or r > ws.max_row:
                continue
            if any(str(ws.cell(row=r, column=c).value or "").strip()
                   for c in range(1, 8)):
                continue
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.border = NO_BORDER
                cell.fill = PatternFill()
            ws.row_dimensions[r].height = 8.0

        ws.freeze_panes = ws.cell(row=hdr + 1, column=3)

    wb.save(args.xlsx)
    print(f"删除 {n_rows} 行(退役模型 + 残留空行),{n_blocks} 行表下内容"
          f"(外部参考 + 冻结协议),{n_cols} 个全空指标列")
    print(f"九个 sheet 的结果表已按统一网格重绘 -> {args.xlsx}")


if __name__ == "__main__":
    main()
