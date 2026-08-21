"""Create workbook sheets for the corpora added to the 12-dataset slate.

    python -m scripts.add_slate_sheets [--xlsx results/PACLock_baseline_matrix_filled.xlsx]

One sheet per new corpus, laid out exactly like the existing ones -- same row
labels, same header strings -- because scripts/fill_xlsx.py locates cells by
matching those strings, so a sheet built to the same shape fills itself once
the runs land. Cells are left EMPTY here: this script builds the frame, never
the numbers.

Column sets follow the corpus's primary metric (metrics.py's PRIMARY_METRIC),
with the primary column first.
"""
from __future__ import annotations

import argparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# corpus -> (sheet name, title, primary metric header, other metric headers)
NEW_SHEETS = {
    "tuep":   ("TUEP", "TUEP — 癫痫 / 非癫痫，记录级二分类（TUAB 的任务形状，TUSZ 的临床问题）",
               "AUROC", ["Balanced Acc", "AUC-PR"]),
    "tuar":   ("TUAR", "TUAR — 伪迹事件形态，3 类（eyem / musc / elec）",
               "Cohen's Kappa", ["Balanced Acc", "Weighted F1"]),
    "adfd":   ("ADFD", "ADFD — 阿尔茨海默 / 额颞叶痴呆 / 健康，3 类被试级诊断",
               "Balanced Acc", ["Cohen's Kappa", "Weighted F1"]),
    "mumtaz": ("Mumtaz2016", "Mumtaz2016 — 抑郁症 / 健康，静息态记录级二分类",
               "AUROC", ["Balanced Acc", "AUC-PR"]),
    "eegmat": ("EEGMat", "EEGMat — 静息 vs 心算，记录级二分类（认知负荷）",
               "AUROC", ["Balanced Acc", "AUC-PR"]),
}

GROUPS = [
    ("A 轻量监督基线", ["SPaRCNet", "ContraWR", "CNN-Transformer", "FFCL",
                        "ST-Transformer"]),
    ("B FM · 官方预训练权重", ["BIOT (pretrained)", "LaBraM-Base (pretrained)",
                               "CBraMod (pretrained)", "EEGPT", "TFM-Tokenizer"]),
    ("C FM · 同 pipeline scratch", ["BIOT (scratch)", "LaBraM-Base (scratch)",
                                    "CBraMod (scratch)",
                                    "PACLock (from scratch, full)",
                                    "PACLock (duplex)", "PACLock (duplex, pretrained)"]),
]
HDR_FILL = PatternFill("solid", fgColor="DDDDDD")


def build(ws, title, primary, others):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "主指标：%s" % primary
    ws["A2"].font = Font(italic=True)
    headers = ["分组", "模型", "参数量 (M)", primary] + others + [
        "Δ%s vs scratch" % primary, "Δ%s vs pt" % primary]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    r = 5
    for group, models in GROUPS:
        for i, m in enumerate(models):
            ws.cell(row=r, column=1, value=group if i == 0 else "")
            ws.cell(row=r, column=2, value=m)
            r += 1
    ws.cell(row=r + 1, column=1,
            value="空单元格 = 尚未跑或未满 3 seed（硬规则 4）；单 seed 数字标注 (1 seed)")
    widths = [24, 30, 12] + [16] * (len(headers) - 3)
    for c, wdt in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = wdt
    return r - 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="results/PACLock_baseline_matrix_filled.xlsx")
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.xlsx)
    for ds, (sheet, title, primary, others) in NEW_SHEETS.items():
        if sheet in wb.sheetnames:
            print("  %-12s sheet exists, left alone" % sheet)
            continue
        # keep the two underscore-prefixed bookkeeping sheets last
        idx = min((wb.sheetnames.index(s) for s in wb.sheetnames
                   if s.startswith("_")), default=len(wb.sheetnames))
        ws = wb.create_sheet(sheet, idx)
        last = build(ws, title, primary, others)
        print("  %-12s created, rows 5-%d" % (sheet, last))
    wb.save(args.xlsx)
    print("saved", args.xlsx)


if __name__ == "__main__":
    main()
