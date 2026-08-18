"""What is in the deliverable xlsx, and what exists in runs/ but is not.

    sbatch slurm/run_cpu.slurm scripts.xlsx_gap

Written before a report, so the answer to "is everything in the table?" is a
list rather than a recollection.
"""
import glob
import json
import os
import statistics as st

from openpyxl import load_workbook

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"
KEY = {"tuab": "auroc", "tuev": "cohen_kappa", "tusz": "pr_auc",
       "chbmit": "pr_auc", "sleepedf": "cohen_kappa", "isruc": "cohen_kappa",
       "physionet_mi": "balanced_acc", "faced": "balanced_acc",
       "bci_iv_2a": "balanced_acc", "tuar": "cohen_kappa"}

wb = load_workbook(XLSX)
print("=== sheets ===")
for ws in wb.worksheets:
    filled = sum(1 for row in ws.iter_rows()
                 for c in row if c.value not in (None, ""))
    print("  %-28s %3d x %-3d  %4d non-empty cells"
          % (ws.title, ws.max_row, ws.max_column, filled))

# every variant that has at least one finished seed
have = {}
for d in sorted(glob.glob("runs/*")):
    base = os.path.basename(d)
    if "-" not in base:
        continue
    ds, variant = base.split("-", 1)
    if ds not in KEY:
        continue
    vals = []
    for f in sorted(glob.glob(d + "/*/result.json")):
        v = json.load(open(f))["test"].get(KEY[ds])
        if v is not None:
            vals.append(v)
    if vals:
        have.setdefault(variant, {})[ds] = (st.mean(vals), len(vals))

# which variant labels does the workbook actually mention anywhere?
text = set()
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                text.add(c.value.strip())
blob = " || ".join(text)

LABEL = {
    "paclock_v2": "PACLock (from scratch, full)",
    "paclock_pt_base": "PACLock (pretrained, base)",
    "paclock_pt_large": "PACLock (pretrained, large)",
    "sparcnet": "SPaRCNet", "contrawr": "ContraWR",
    "cnn_transformer": "CNN-Transformer", "ffcl": "FFCL",
    "st_transformer": "ST-Transformer", "biot_prest16": "BIOT (pretrained)",
    "labram_pretrained": "LaBraM-Base (pretrained)",
    "cbramod_pretrained": "CBraMod (pretrained)", "eegpt_pretrained": "EEGPT",
    "tfm_pretrained": "TFM-Tokenizer", "biot_scratch": "BIOT (scratch)",
    "cbramod_scratch": "CBraMod (scratch)",
    "labram_scratch": "LaBraM-Base (scratch)",
}

print("\n=== variants with finished runs, and whether the workbook has a row ===")
print("%-30s %-8s %-40s %s" % ("variant", "corpora", "row label in workbook", "3-seed corpora"))
for v in sorted(have, key=lambda k: -len(have[k])):
    lab = LABEL.get(v)
    if lab and lab in blob:
        state = "yes  (%s)" % lab
    elif lab:
        state = "LABEL KNOWN, NOT IN SHEET (%s)" % lab
    else:
        state = "** NO ROW -- not in the table **"
    n3 = sum(1 for _, (_, n) in have[v].items() if n >= 3)
    print("%-30s %-8d %-40s %d" % (v, len(have[v]), state[:40], n3))

print("\n=== our variants that are NOT represented at all ===")
for v in sorted(have):
    if not v.startswith("paclock"):
        continue
    if LABEL.get(v) and LABEL[v] in blob:
        continue
    cells = ", ".join("%s %.4f(%d)" % (ds, m, n)
                      for ds, (m, n) in sorted(have[v].items()))
    print("  %-28s %s" % (v, cells))
