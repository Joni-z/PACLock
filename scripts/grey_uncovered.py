"""Grey out matrix cells whose baseline never ran that corpus in its own paper.

Why this matters: the benchmark's rule is that every baseline runs its own
repo's recipe, which makes a cell a faithful reproduction of published
practice -- but only where the baseline's authors actually evaluated that
corpus. Where they did not, no official recipe exists for it, and the
hyperparameters were carried over from a corpus they did run. Those cells are
still legitimate numbers, but they measure "this architecture under a recipe
we chose for it", not "this baseline's published capability", and they are
where a baseline can silently fail (CNN-Transformer collapses to chance on
FACED and PhysioNet-MI at BIOT's lr=1e-3, which BIOT never applied to either
corpus).

Coverage is taken from each paper's own evaluation table, read directly:
  BIOT (Yang 2023, Table 1)            CHB-MIT, IIIC, TUAB, TUEV (+PTB-XL, HAR)
  LaBraM (Jiang 2024)                  TUAB, TUEV main; SEED-V/MoBI appendix
  CBraMod (Wang 2025, Table 1)         FACED, SEED-V, PhysioNet-MI, SHU-MI,
                                       ISRUC, CHB-MIT, BCIC2020-3, Mumtaz2016,
                                       SEED-VIG, MentalArithmetic, TUEV, TUAB
  EEGPT (Wang 2024, Table 1)           BCIC-2A, BCIC-2B, Sleep-EDFx, KaggleERN,
                                       PhysioP300, TUAB, TUEV
  TFM-Tokenizer (2026)                 TUEV, TUAB, CHB-MIT, IIIC

Group A (SPaRCNet, ContraWR, CNN-Transformer, FFCL, ST-Transformer) enter this
benchmark through BIOT's implementation and recipe, so they inherit BIOT's
coverage rather than their own original papers'.

PACLock rows are never greyed: this is our model and every corpus here is one
we target deliberately.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"

SHEET_TO_CORPUS = {
    "TUAB": "tuab", "TUEV": "tuev", "TUSZ": "tusz", "CHB-MIT": "chbmit",
    "Sleep-EDF": "sleepedf", "ISRUC": "isruc", "PhysioNet-MI": "physionet_mi",
    "FACED": "faced", "BCI-IV-2a": "bci_iv_2a",
}

BIOT_COV = {"chbmit", "tuab", "tuev"}
COVERAGE = {
    "SPaRCNet": BIOT_COV,
    "ContraWR": BIOT_COV,
    "CNN-Transformer": BIOT_COV,
    "FFCL": BIOT_COV,
    "ST-Transformer": BIOT_COV,
    "BIOT (pretrained)": BIOT_COV,
    "BIOT (scratch)": BIOT_COV,
    "LaBraM-Base (pretrained)": {"tuab", "tuev"},
    "LaBraM-Base (scratch)": {"tuab", "tuev"},
    "CBraMod (pretrained)": {"faced", "physionet_mi", "isruc", "chbmit", "tuev", "tuab"},
    "CBraMod (scratch)": {"faced", "physionet_mi", "isruc", "chbmit", "tuev", "tuab"},
    "EEGPT": {"bci_iv_2a", "sleepedf", "tuab", "tuev"},
    "TFM-Tokenizer": {"tuev", "tuab", "chbmit"},
}

GREY_FONT = Font(color="909090", italic=True)
GREY_FILL = PatternFill("solid", fgColor="EFEFEF")

wb = openpyxl.load_workbook(XLSX)
n_grey = n_kept = 0
for sheet, corpus in SHEET_TO_CORPUS.items():
    ws = wb[sheet]
    hdr = 4
    ncol = ws.max_column
    for r in range(hdr + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        cov = COVERAGE.get(label)
        if cov is None:                       # PACLock rows and anything unlisted
            n_kept += 1
            continue
        if corpus in cov:
            n_kept += 1
            continue
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = GREY_FONT
            cell.fill = GREY_FILL
        n_grey += 1

# legend under each sheet's table so the grey is self-explanatory in the file
for sheet in SHEET_TO_CORPUS:
    ws = wb[sheet]
    row = ws.max_row + 2
    note = ("灰色行 = 该 baseline 的原论文未评测本数据集,不存在官方超参;"
            "所用配方由我们从该模型覆盖过的数据集迁移而来。"
            "这些数字反映的是「该架构在我们所选配方下的表现」,"
            "不是「该 baseline 已发表的能力」。")
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(size=9, color="808080", italic=True)

wb.save(XLSX)
print("greyed %d cells, left %d untouched" % (n_grey, n_kept))
