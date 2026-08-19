"""Insert the plain-hybrid row into each per-corpus main sheet.

    sbatch slurm/run_cpu.slurm scripts.add_hybrid_row

Until now our variants lived only in the ablation sheet and the nine main
sheets carried exactly three PACLock rows (scratch / pt-base / pt-large).
Zhizhe asked for hybrid in the main sheets, so it goes in as a fourth row
directly under pt-large, with the same column conventions the sheet already
uses:

  * metric cells come from runs/<ds>-paclock_hybrid/*/result.json; a cell is
    written as mean±std when it has >=2 seeds and as "value (1 seed)" in
    italic when it has one -- the seed count is never dropped, because a
    single-seed number silently formatted like a 3-seed one is the one way
    this row could mislead;
  * the three delta columns follow the sheet's own convention (this row's
    primary-metric mean minus the scratch / pt-base / pt-large rows'), read
    back from the sheet itself so the arithmetic cannot diverge from what the
    sheet displays;
  * corpora with no hybrid run get the row with an em-dash, so "not run" is
    visible rather than ambiguous.

Styling is copied cell-by-cell from the pt-large row above.
"""
import glob
import json
import statistics as st
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"
SHEETS = {"TUAB": "tuab", "TUEV": "tuev", "TUSZ": "tusz", "CHB-MIT": "chbmit",
          "Sleep-EDF": "sleepedf", "ISRUC": "isruc",
          "PhysioNet-MI": "physionet_mi", "FACED": "faced",
          "BCI-IV-2a": "bci_iv_2a"}
HEADER_TO_KEY = {"Balanced Acc": "balanced_acc", "AUC-PR": "pr_auc",
                 "AUROC": "auroc", "Cohen's Kappa": "cohen_kappa",
                 "Kappa": "cohen_kappa", "Weighted F1": "weighted_f1",
                 "F1 (weighted)": "weighted_f1"}
LABEL = "PACLock (hybrid)"


def seed_values(ds):
    rows = []
    n_params = None
    for f in sorted(glob.glob(f"runs/{ds}-paclock_hybrid/*/result.json")):
        r = json.load(open(f))
        rows.append(r["test"])
        n_params = r.get("n_params_M", n_params)
    return rows, n_params


def fmt(vals):
    if not vals:
        return None
    if len(vals) == 1:
        return "%.4f (1 seed)" % vals[0]
    return "%.4f±%.4f (%d)" % (st.mean(vals), st.stdev(vals), len(vals))


def mean_from_cell(text):
    """Leading float out of 'mean±std' / 'x (1 seed)' cell text."""
    if not text:
        return None
    tok = str(text).split("±")[0].split("(")[0].strip()
    try:
        return float(tok)
    except ValueError:
        return None


def main():
    wb = load_workbook(XLSX)
    for sheet, ds in SHEETS.items():
        ws = wb[sheet]
        hdr_row = next(r for r in range(1, 12)
                       if ws.cell(r, 1).value == "分组")
        headers = {c: (ws.cell(hdr_row, c).value or "")
                   for c in range(1, ws.max_column + 1)}

        def find_row(substr):
            for r in range(hdr_row + 1, ws.max_row + 1):
                if substr in str(ws.cell(r, 2).value or ""):
                    return r
            return None

        if find_row("hybrid"):
            print(f"{sheet}: hybrid row already present, skipping")
            continue
        anchor = find_row("(pretrained, large)")
        scratch_r = find_row("(from scratch")
        base_r = find_row("(pretrained, base)")
        assert anchor and scratch_r and base_r, sheet

        new_r = anchor + 1
        ws.insert_rows(new_r)
        for c in range(1, ws.max_column + 1):
            src = ws.cell(anchor, c)
            dst = ws.cell(new_r, c)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)

        rows, n_params = seed_values(ds)
        ws.cell(new_r, 2, LABEL)
        if n_params:
            ws.cell(new_r, 3, round(n_params, 2))

        prim_key = None
        prim_col = None
        for c, h in headers.items():
            if h.startswith("Δ"):
                continue
            key = HEADER_TO_KEY.get(h.strip())
            if not key:
                continue
            vals = [t[key] for t in rows if key in t]
            cell = ws.cell(new_r, c)
            cell.value = fmt(vals) if vals else "—"
            if len(vals) == 1:
                f = copy(cell.font)
                f.italic = True
                cell.font = f
        # primary metric = the metric named inside the delta headers
        for c, h in headers.items():
            if h.startswith("Δ") and "scratch" in h:
                name = h[1:].split(" vs ")[0].strip()
                prim_key = HEADER_TO_KEY.get(name)
                prim_col = next(cc for cc, hh in headers.items()
                                if hh.strip() == name)
                break
        if prim_key and rows:
            mine = st.mean([t[prim_key] for t in rows if prim_key in t])
            for ref_row, tag in ((scratch_r, "scratch"), (base_r, "pt-base"),
                                 (anchor, "pt-large")):
                ref = mean_from_cell(ws.cell(ref_row, prim_col).value)
                col = next((cc for cc, hh in headers.items()
                            if hh.startswith("Δ") and tag in hh), None)
                if col and ref is not None:
                    cell = ws.cell(new_r, col, "%+.4f" % (mine - ref))
                    if len(rows) == 1:
                        f = copy(cell.font)
                        f.italic = True
                        cell.font = f
        n = len(rows)
        print(f"{sheet}: row {new_r}, {n} seed(s)"
              + ("" if n else " -- placeholder only"))
    wb.save(XLSX)
    print("saved")


if __name__ == "__main__":
    main()
