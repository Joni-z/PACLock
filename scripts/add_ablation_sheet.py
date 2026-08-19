"""Rebuild the ablation sheet in the deliverable workbook.

    sbatch slurm/run_cpu.slurm scripts.add_ablation_sheet

The nine per-corpus sheets are the frozen baseline matrix and are not touched.
This sheet is every PACLock variant that has a finished run, including TUAR
(added later, not part of the nine) and the TUEV sample-size ladder.

Hard rule 4 (three seeds or the cell stays empty) governs what may enter the
PAPER table. It does not govern this sheet, which exists to show what has been
measured. Every cell therefore carries its seed count in brackets and one-seed
cells are written normally -- but the count is never dropped, because a
one-seed number read as a three-seed number is the one way this sheet could
mislead.
"""
import glob
import json
import os
import statistics as st

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"
SHEET = "_消融"

DS = ["tuab", "tuev", "tusz", "chbmit", "sleepedf", "isruc",
      "physionet_mi", "faced", "bci_iv_2a", "tuar"]
DS_LABEL = {"tuab": "TUAB", "tuev": "TUEV", "tusz": "TUSZ", "chbmit": "CHB-MIT",
            "sleepedf": "Sleep-EDF", "isruc": "ISRUC",
            "physionet_mi": "PhysioNet-MI", "faced": "FACED",
            "bci_iv_2a": "BCI-IV-2a", "tuar": "TUAR"}
KEY = {"tuab": "auroc", "tuev": "cohen_kappa", "tusz": "pr_auc",
       "chbmit": "pr_auc", "sleepedf": "cohen_kappa", "isruc": "cohen_kappa",
       "physionet_mi": "balanced_acc", "faced": "balanced_acc",
       "bci_iv_2a": "balanced_acc", "tuar": "cohen_kappa"}

ROWS = [
    ("参照", "paclock_v2", "PACLock v2(冻结配置:pac_interaction + product)"),

    ("tokenizer 消融", "paclock_rawtok", "raw tokenizer(去掉整个交互)"),
    ("tokenizer 消融", "paclock_raw", "raw tokenizer(TUAR 专跑)"),
    ("tokenizer 消融", "paclock_concat", "concat 融合(SleepPACNet 式)"),
    ("tokenizer 消融", "paclock_pac_uniform", "uniform 耦合(去掉测得的 α 与 ∠Z)"),
    ("tokenizer 消融", "paclock_pac", "PAC 协议预处理(0.5Hz 高通、无 notch)"),

    ("架构改动", "paclock_hybrid", "hybrid(raw 行 + 交互行并列)"),
    ("架构改动", "paclock_hybrid_gate", "hybrid + 逐频带可学门"),
    ("架构改动", "paclock_hybrid_attn", "hybrid + attn 头"),
    ("架构改动", "paclock_hybrid_sp", "hybrid + spatial 头"),
    ("架构改动", "paclock_hybrid_sp_gate", "hybrid + spatial 头 + 门"),
    ("架构改动", "paclock_fuse", "fused:行内融合 blend(β 零初始化)"),
    ("架构改动", "paclock_fusegate", "fused:行内融合 gated(内容门)"),
    ("架构改动", "paclock_rot2", "rotation(耦合只旋转不缩放)"),
    ("架构改动", "paclock_rot", "rotation(跑在 PAC 协议数据上,与上行不可比)"),
    ("架构改动", "paclock_sphead", "spatial 头(保留电极身份)"),
    ("架构改动", "paclock_rot_sphead", "rotation + spatial 头"),

    ("预训练", "paclock_pt_base", "预训练 base(60k)"),
    ("预训练", "paclock_pt_large", "预训练 large(60k)"),
    ("预训练", "paclock_rawpt_large", "raw tokenizer 预训练 large(60k)"),
    ("预训练", "paclock_pt_excl", "预训练池剔除 TUSZ/CHB-MIT"),
    ("预训练", "paclock_pt_base_p200", "预训练 base,微调 patch_len=200"),
    ("预训练", "paclock_ft_base", "早期微调配方(6k checkpoint)"),
    ("预训练", "paclock_ft_base_v2", "早期微调配方 v2(6k checkpoint)"),
    ("预训练", "paclock_ft_large", "早期微调配方 large(6k checkpoint)"),

    ("容量阶梯", "paclock_raw_tiny", "raw,d_model 32"),
    ("容量阶梯", "paclock_raw_small", "raw,d_model 64"),
    ("容量阶梯", "paclock_raw_large", "raw,d_model 256"),
    ("容量阶梯", "paclock_raw_wide", "raw,加宽"),
    ("容量阶梯", "paclock_raw_nb16", "raw,16 频带"),
    ("容量阶梯", "paclock_raw_p100", "raw,patch_len 100"),
    ("容量阶梯", "paclock_raw_drop5", "raw,dropout 0.5"),
    ("容量阶梯", "paclock_raw_headattn", "raw,attn 头"),

    ("已否决", "paclock_xyz", "spatial_pe: xyz(蒙太奇坐标)"),
    ("已否决", "paclock_wn", "每窗口归一化"),
    ("已否决", "paclock_wn_raw", "每窗口归一化 + raw"),
    ("已否决", "paclock_sn", "每受试者归一化"),
    ("已否决", "paclock_sn_raw", "每受试者归一化 + raw"),
    ("已否决", "paclock_sn_attn_raw", "每受试者归一化 + raw + attn 头"),
    ("已否决", "paclock_clean", "FACED 伪迹清洗"),
    ("已否决", "paclock_lr1e3", "lr 1e-3"),
    ("已否决", "paclock_lr3e4", "lr 3e-4"),

    ("早期 pilot", "paclock_pilot_frozen", "冻结前端 pilot"),
    ("早期 pilot", "paclock_pilot_unfiltered", "未滤波 pilot"),
]

# TUEV sample-size ladder: a separate question, so a separate little block
SIZE_ROWS = [("paclock_v2", "PACLock v2"), ("sparcnet", "SPaRCNet")]
CAPS = [("_n2160", "n=2160"), ("_n6720", "n=6720"),
        ("_n20000", "n=20000"), ("", "全量 68445")]

HDR = PatternFill("solid", fgColor="D9D9D9")
GRP = PatternFill("solid", fgColor="EDF3FA")
ONE = PatternFill("solid", fgColor="FFF7E6")      # single seed: noted, not excluded
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def cell(ds, variant):
    vals = []
    for p in sorted(glob.glob("runs/%s-%s/*/result.json" % (ds, variant))):
        v = json.load(open(p))["test"].get(KEY[ds])
        if v is not None:
            vals.append(v)
    if not vals:
        return None, 0
    s = st.stdev(vals) if len(vals) > 1 else 0.0
    return "%.4f±%.4f" % (st.mean(vals), s), len(vals)


def put(ws, r, c, txt, n):
    x = ws.cell(r, c)
    x.border, x.alignment = BOX, CENTER
    if txt is None:
        return 0, 0
    x.value = "%s (%d)" % (txt, n)
    if n < 3:
        x.fill = ONE
    return 1, (1 if n < 3 else 0)


def main():
    wb = load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    ws["A1"] = "PACLock 自身消融 —— 与冻结的 baseline 主表分开,主表未改动"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("每格 mean±std,括号内为 seed 数。淡黄底 = 只有 1-2 个 seed,"
                "数字可用但不可当作 3-seed 结论;硬规则 4 只约束论文正表,不约束本表。"
                "指标:TUAB=AUROC,TUSZ/CHB-MIT=PR-AUC,TUEV/Sleep-EDF/ISRUC/TUAR=kappa,"
                "其余=balanced acc。空格 = 未跑或在跑。")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 32

    hdr = 4
    for col, name in ((1, "分组"), (2, "变体")):
        c = ws.cell(hdr, col, name)
        c.fill, c.font, c.border = HDR, Font(bold=True), BOX
    for j, ds in enumerate(DS):
        c = ws.cell(hdr, 3 + j, DS_LABEL[ds])
        c.fill, c.font, c.alignment, c.border = HDR, Font(bold=True), CENTER, BOX

    r, last, n_all, n_one = hdr + 1, None, 0, 0
    for group, variant, label in ROWS:
        if not any(cell(ds, variant)[0] for ds in DS):
            continue
        g = ws.cell(r, 1, group if group != last else "")
        if group != last:
            g.font, g.fill = Font(bold=True), GRP
        last = group
        g.border = BOX
        ws.cell(r, 2, label).border = BOX
        for j, ds in enumerate(DS):
            txt, n = cell(ds, variant)
            a, b = put(ws, r, 3 + j, txt, n)
            n_all += a
            n_one += b
        r += 1

    # --- sample-size ladder, TUEV only ---
    r += 1
    ws.cell(r, 1, "样本量消融").font = Font(bold=True)
    ws.cell(r, 1).fill = GRP
    ws.cell(r, 2, "TUEV 训练集降采样(kappa) —— 检验等级三是否只是样本量问题").font = Font(bold=True)
    r += 1
    ws.cell(r, 2, "模型").font = Font(bold=True)
    ws.cell(r, 2).fill = HDR
    ws.cell(r, 2).border = BOX
    for j, (_, lab) in enumerate(CAPS):
        c = ws.cell(r, 3 + j, lab)
        c.fill, c.font, c.alignment, c.border = HDR, Font(bold=True), CENTER, BOX
    r += 1
    for variant, label in SIZE_ROWS:
        ws.cell(r, 2, label).border = BOX
        for j, (suf, _) in enumerate(CAPS):
            txt, n = cell("tuev", variant + suf)
            a, b = put(ws, r, 3 + j, txt, n)
            n_all += a
            n_one += b
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    for j in range(len(DS)):
        ws.column_dimensions[chr(ord("C") + j)].width = 18
    ws.freeze_panes = "C5"

    wb.save(XLSX)
    print("sheet %r: %d cells, %d of them at 1-2 seeds (cream, not excluded)"
          % (SHEET, n_all, n_one))


if __name__ == "__main__":
    main()
