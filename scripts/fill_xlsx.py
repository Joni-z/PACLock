"""Write group-A results into a COPY of the matrix workbook.

    python -m scripts.fill_xlsx --xlsx PACLock_baseline_matrix.xlsx \
                                [--out PACLock_baseline_matrix_filled.xlsx]

Never touches the input file. The workbook is the source of truth for the frozen
protocols, and a script that rewrites it in place can damage formatting, notes or
formulas that are not reproducible from ``runs/``.

Cells are filled only where the result is admissible:

* hard rule 4 -- fewer than 3 seeds leaves the cell empty
* hard rule 3 -- any seed flagged mis-configured leaves the cell empty and adds a
  note; the protocol says such a cell is "拒绝写入", so writing the number even
  greyed out would overstate it

Every filled cell gets ``mean ± std`` over 3 seeds. A ``_notes`` sheet records,
per cell, the seed values, the admissibility verdict and the reason -- so a
reader can always see what was withheld and why.
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import shutil
from collections import defaultdict

import numpy as np
import openpyxl

MODEL_ROW_LABEL = {
    # group A -- light supervised baselines
    "sparcnet": "SPaRCNet",
    "contrawr": "ContraWR",
    "cnn_transformer": "CNN-Transformer",
    "ffcl": "FFCL",
    "st_transformer": "ST-Transformer",
    # group B -- foundation models with their official pretrained weights
    "biot_prest16": "BIOT (pretrained)",
    "labram_pretrained": "LaBraM-Base (pretrained)",
    "cbramod_pretrained": "CBraMod (pretrained)",
    "eegpt_pretrained": "EEGPT",
    "tfm_pretrained": "TFM-Tokenizer",
    # group C -- the same architectures trained from scratch
    "biot_scratch": "BIOT (scratch)",
    "cbramod_scratch": "CBraMod (scratch)",
    "labram_scratch": "LaBraM-Base (scratch)",
    # ours, in the same block -- the run directory is <ds>-paclock_full, so the
    # variant key is paclock_full. Its absence here was silent: the cell simply
    # reported "no row label" in the notes sheet and the row stayed blank, which
    # is indistinguishable from "the runs are not finished yet".
    "paclock_full": "PACLock (from scratch, full)",
}

# Rows the workbook does not carry, so a run for them has nowhere to go.
# Listed explicitly rather than silently skipped, since "no row" and "we forgot
# to map the label" look identical in the output otherwise.
NO_ROW_IN_WORKBOOK = {
    "eegpt_scratch", "tfm_scratch",
    # group D: protocol-variant and pilot arms. They belong in the sensitivity
    # analysis, not the main matrix, so having no row is correct -- but they must
    # be listed, because "deliberately has no row" and "we forgot to map the
    # label" produce the same silent blank. That is exactly how paclock_full
    # stayed empty while its runs sat finished on disk.
    "cbramod_pac", "paclock_pac",
    "paclock_pilot_frozen", "paclock_pilot_unfiltered",
    # architecture search + tokenisation ablations. These are not matrix cells:
    # `cand_*` are single-seed screening arms for the PACLock structure search,
    # `biot_tok100`/`biot_hop50` are the token-density control that asks whether
    # a shorter patch helps a non-PAC tokeniser too. Both belong in the ablation
    # write-up, not in a benchmark row.
    "biot_tok100", "biot_hop50",
}


def _no_row(model: str) -> bool:
    return model in NO_ROW_IN_WORKBOOK or model.startswith("cand_")
DATASET_TO_SHEET = {
    "tuab": "TUAB", "tuev": "TUEV", "tusz": "TUSZ", "chbmit": "CHB-MIT",
    "sleepedf": "Sleep-EDF", "isruc": "ISRUC", "physionet_mi": "PhysioNet-MI",
    "faced": "FACED", "bci_iv_2a": "BCI-IV-2a",
}
# our metric key -> the column header used in each sheet's upper (results) block
METRIC_HEADER = {
    "balanced_acc": "Balanced Acc",
    "cohen_kappa": "Cohen's Kappa",
    "weighted_f1": "Weighted F1",
    "auroc": "AUROC",
    "pr_auc": "AUC-PR",
}


def load_runs(runs_dir: str) -> dict:
    """Group results by (dataset, variant).

    The variant comes from the run directory (``runs/tuab-biot_prest16/``), not
    from ``result.json``'s ``model`` field. That field holds the architecture
    ("biot"), which cannot tell the pretrained row apart from the scratch one --
    they are different rows of the workbook and must never be merged.
    """
    out = defaultdict(list)
    for f in sorted(glob.glob(os.path.join(runs_dir, "*", "seed*", "result.json"))):
        with open(f) as fh:
            r = json.load(fh)
        run_name = os.path.basename(os.path.dirname(os.path.dirname(f)))
        ds = r["dataset"]
        variant = run_name[len(ds) + 1:] if run_name.startswith(ds + "-") else r["model"]
        out[(ds, variant)].append(r)
    return out


# Settings that change what a number means, recorded per cell so the comparison
# can be checked from the workbook alone. Hard rule 2 has every model running its
# own repo's recipe, which is the right call but means the cells in one column
# are NOT under a shared schedule -- and a reader who does not know that will
# read the table wrong. docs/RECIPE_AUDIT.md gives the justification for each;
# this is the per-cell record of what was actually used.
RECIPE_KEYS = ("lr", "batch_size", "epochs", "patience", "eval_every_steps",
               "loss", "label_smoothing", "grad_clip", "select_metric",
               "multi_lr", "loader_divisor")


def recipe_fingerprint(run: dict) -> str:
    cfg = run.get("config") or {}
    parts = []
    # Surfaced first because it qualifies the number: a run stopped by the
    # wall-clock budget did not finish its own schedule.
    if run.get("stopped_by") and run["stopped_by"] != "epochs":
        parts.append(f"stopped_by={run['stopped_by']}")
    for k in RECIPE_KEYS:
        if k in cfg and cfg[k] is not None:
            parts.append(f"{k}={cfg[k]}")
    mk = cfg.get("model_kwargs") or {}
    # PACLock's architecture of record lives in model_kwargs, and two of those
    # keys silently fall back to a different architecture when omitted.
    for k in ("tokenizer_mode", "pac_token_mode", "interaction_mode",
              "band_pe", "spatial_pe", "input_norm"):
        if k in mk:
            parts.append(f"{k}={mk[k]}")
    return " ".join(parts)


def find_results_block(ws):
    """Locate the header row of the upper results block and its column map.

    The sheets put the model matrix above a '外部参考值' block that repeats the
    same metric names, so the search stops at the first header row -- matching
    the second one would write our numbers into the published-anchor table.
    """
    for row in ws.iter_rows(min_row=1, max_row=12):
        vals = [c.value for c in row]
        if "模型" in vals:
            cols = {}
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in METRIC_HEADER.values():
                    cols[c.value.strip()] = c.column
            model_col = vals.index("模型") + 1
            param_col = next(
                (c.column for c in row
                 if isinstance(c.value, str) and "参数量" in c.value), None)
            return row[0].row, model_col, cols, param_col
    return None, None, None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default=None)
    ap.add_argument("--runs", default="runs")
    args = ap.parse_args()

    out_path = args.out or args.xlsx.replace(".xlsx", "_filled.xlsx")
    if os.path.abspath(out_path) == os.path.abspath(args.xlsx):
        raise SystemExit("refusing to overwrite the input workbook")
    # Copy only if the output does not exist yet. add_faced_sheet runs first and
    # writes the same file; re-copying the input here would silently throw its
    # new FACED sheet away and every FACED cell would be reported as "skipped".
    if not os.path.exists(out_path):
        shutil.copy2(args.xlsx, out_path)
        print(f"created {out_path} from {args.xlsx}")
    else:
        print(f"appending to existing {out_path}")
    print(f"input  : {args.xlsx} (untouched)")
    print(f"output : {out_path}")

    runs = load_runs(args.runs)
    wb = openpyxl.load_workbook(out_path)
    notes: list[list] = []
    n_filled = n_withheld = 0

    for (ds, model), rs in sorted(runs.items()):
        if _no_row(model):
            notes.append([ds, model, "-", "skipped",
                          "the workbook has no row for this variant"])
            continue
        sheet = DATASET_TO_SHEET.get(ds)
        if sheet is None or sheet not in wb.sheetnames:
            notes.append([ds, model, "-", "skipped", f"no sheet for dataset {ds!r}"])
            continue
        ws = wb[sheet]
        hdr_row, model_col, cols, param_col = find_results_block(ws)
        if hdr_row is None:
            notes.append([ds, model, "-", "skipped", "results header row not found"])
            continue

        if model not in MODEL_ROW_LABEL:
            notes.append([ds, model, "-", "skipped", f"no row label for {model!r}"])
            continue
        label = MODEL_ROW_LABEL[model]
        # only the scratch/pretrained-free group-A rows: take the FIRST match
        # below the header, which is the A block
        target = None
        for r in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=model_col).value
            if isinstance(v, str) and v.strip() == label:
                target = r
                break
        if target is None:
            notes.append([ds, model, "-", "skipped", f"row {label!r} not found"])
            continue

        key = rs[0]["primary_metric"]
        seeds = sorted(r["seed"] for r in rs)
        bad = sorted(r["seed"] for r in rs if not r["verdict"]["ok"])
        vals_primary = [r["test"][key] for r in rs]

        if len(rs) < 3:
            n_withheld += 1
            notes.append([ds, model, str(seeds), "withheld (rule 4)",
                          f"only {len(rs)} seeds; primary {key} "
                          f"{np.mean(vals_primary):.4f}"])
            continue
        if bad:
            n_withheld += 1
            reasons = {r["seed"]: r["verdict"]["reason"] for r in rs if not r["verdict"]["ok"]}
            notes.append([ds, model, str(seeds), "withheld (rule 3)",
                          f"mis-configured seeds {bad}: {reasons}; measured "
                          f"{key} {np.mean(vals_primary):.4f}±"
                          f"{np.std(vals_primary):.4f}"])
            continue

        wrote = []
        for mkey, header in METRIC_HEADER.items():
            if mkey not in rs[0]["test"] or header not in cols:
                continue
            v = [r["test"][mkey] for r in rs]
            ws.cell(row=target, column=cols[header],
                    value=f"{np.mean(v):.4f}±{np.std(v):.4f}")
            wrote.append(header)
        n_filled += 1
        spread = np.std(vals_primary) / abs(np.mean(vals_primary)) if np.mean(vals_primary) else 0
        # Overwrite the nominal parameter count with what the run actually
        # built. The nominal figures were inconsistent in three ways: EEGPT's
        # 4.7 refers to a variant that was never released (the only public
        # checkpoint is large4E, 25.7M); CBraMod's 4 counts the backbone alone
        # while its all_patch_reps head scales with channels x patches, giving
        # 9.9M on Sleep-EDF and 56.3M on FACED; and SPaRCNet's 0.79 is 26% under
        # what its official implementation instantiates here. Reporting the
        # measured value per cell removes all three, and removes the need for a
        # footnote explaining which convention each row follows.
        meas = rs[0].get("n_params_M")
        if param_col and isinstance(meas, (int, float)):
            ws.cell(row=target, column=param_col, value=round(meas, 2))
        pstr = f"  [实测 {meas:.2f}M]" if isinstance(meas, (int, float)) else ""
        notes.append([ds, model, str(seeds), "filled",
                      pstr.strip() + " " + f"{key} per-seed " +
                      ", ".join(f"{s}:{r['test'][key]:.4f}" for s, r in
                                zip(seeds, sorted(rs, key=lambda x: x['seed']))) +
                      (f"  [seed spread {spread:.0%} -- mean is a weak summary]"
                       if spread > 0.15 else ""),
                      recipe_fingerprint(rs[0])])

    ns = wb.create_sheet("_填写记录") if "_填写记录" not in wb.sheetnames \
        else wb["_填写记录"]
    ns.delete_rows(1, ns.max_row)
    ns.append(["dataset", "model", "seeds", "状态", "说明", "recipe"])
    for r in notes:
        ns.append(r)

    wb.save(out_path)
    print(f"\n填入 {n_filled} 个单元格,按规则留空 {n_withheld} 个")
    print("每个单元格的 seed 明细与留空原因见 '_填写记录' sheet")
    for r in notes:
        if r[3] != "filled":
            print(f"  留空: {r[0]}/{r[1]} -- {r[3]}")


if __name__ == "__main__":
    main()
