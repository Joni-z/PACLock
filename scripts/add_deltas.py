"""Fill the "Delta vs PACLock" column that every sheet's template carries but
fill_xlsx.py never populated.

    python -m scripts.add_deltas --xlsx results/PACLock_baseline_matrix_filled.xlsx

Each sheet's results header ends in a column titled e.g. "ΔCohen's Kappa vs
PACLock" -- the metric name in the middle is the sheet's own primary metric, so
the target column is found by matching it against the same METRIC_HEADER
columns fill_xlsx.py already writes, not by hardcoding which metric goes with
which sheet.

Requires the "PACLock (from scratch, full)" row to be filled on that sheet: a
sheet where PACLock's own cell is empty (three of the nine, at the time this
was written -- the v2 sweep had not reached them) has no reference to diff
against, so its delta column is correctly left blank rather than computed
against a stale or missing number.

Modifies the workbook in place, run after fill_xlsx.py and before
normalize_xlsx.py.
"""

from __future__ import annotations

import argparse
import re

import openpyxl

from scripts.fill_xlsx import METRIC_HEADER, MODEL_ROW_LABEL, find_results_block

PACLOCK_LABEL = MODEL_ROW_LABEL["paclock_v2"]
VALUE_RE = re.compile(r"^(-?\d+\.\d+)")


def parse_mean(cell_value) -> float | None:
    if not isinstance(cell_value, str):
        return None
    m = VALUE_RE.match(cell_value.strip())
    return float(m.group(1)) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    n_sheets_done = n_sheets_skipped = n_cells = 0

    for name in wb.sheetnames:
        if name in ("README", "_填写记录"):
            continue
        ws = wb[name]
        hdr_row, model_col, cols, _ = find_results_block(ws)
        if hdr_row is None:
            continue

        delta_col = next(
            (c.column for c in ws[hdr_row]
             if isinstance(c.value, str) and c.value.strip().startswith("Δ")),
            None)
        if delta_col is None:
            continue

        delta_header = ws.cell(row=hdr_row, column=delta_col).value.strip()
        # "ΔCohen's Kappa vs PACLock" -> "Cohen's Kappa"; matched against the
        # same header text fill_xlsx.py used, so this can't silently pick the
        # wrong metric if a sheet's header text ever changes.
        metric_name = delta_header.lstrip("Δ").removesuffix(" vs PACLock").strip()
        metric_col = cols.get(metric_name)
        if metric_col is None:
            raise SystemExit(
                f"{name}: delta header {delta_header!r} names metric "
                f"{metric_name!r}, which is not one of this sheet's columns "
                f"{list(cols)} -- fix the template or METRIC_HEADER"
            )

        paclock_row = None
        for r in range(hdr_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=model_col).value
            if isinstance(v, str) and v.strip() == PACLOCK_LABEL:
                paclock_row = r
                break
        if paclock_row is None:
            n_sheets_skipped += 1
            print(f"  {name}: no PACLock row filled yet -- delta column left blank")
            continue
        ref = parse_mean(ws.cell(row=paclock_row, column=metric_col).value)
        if ref is None:
            n_sheets_skipped += 1
            print(f"  {name}: PACLock row present but {metric_name} cell is empty")
            continue

        filled_here = 0
        for r in range(hdr_row + 1, ws.max_row + 1):
            if r == paclock_row:
                continue
            label = ws.cell(row=r, column=model_col).value
            if not isinstance(label, str) or not label.strip():
                continue
            v = parse_mean(ws.cell(row=r, column=metric_col).value)
            if v is None:
                continue
            ws.cell(row=r, column=delta_col, value=f"{v - ref:+.4f}")
            filled_here += 1

        n_sheets_done += 1
        n_cells += filled_here
        print(f"  {name}: PACLock {metric_name}={ref:.4f} (row {paclock_row}) "
              f"-> {filled_here} deltas written")

    wb.save(args.xlsx)
    print(f"\n{n_cells} delta cells written across {n_sheets_done} sheets "
          f"({n_sheets_skipped} skipped: no PACLock reference yet)")


if __name__ == "__main__":
    main()
