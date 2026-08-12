"""Add the FACED sheet to a COPY of the workbook.

    python -m scripts.add_faced_sheet --xlsx <in.xlsx> [--out <out.xlsx>]

The current workbook has no FACED sheet: FACED was in the first version, then
replaced by SEED-V, and SEED-V has now been replaced back by FACED. Rather than
inventing a protocol, this reproduces the frozen FACED protocol from the first
version verbatim (it is also transcribed in docs/PROTOCOLS.md sec.8) and lays the
sheet out exactly like the others, so scripts/fill_xlsx.py finds the same header
row and column positions.

Never edits the input file.
"""

from __future__ import annotations

import argparse
import os
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font

TITLE = "FACED — 情绪识别(9 类)"
PRIMARY = "主指标:Balanced Accuracy"
HEADER = ["分组", "模型", "参数量 (M)", "Balanced Acc", "Cohen's Kappa",
          "Weighted F1", "ΔBalanced Acc vs PACLock"]

ROWS = [
    ("A 轻量监督基线", "SPaRCNet", 0.79),
    (None, "ContraWR", 1.6),
    (None, "CNN-Transformer", 3.2),
    (None, "FFCL", 2.4),
    (None, "ST-Transformer", 3.5),
    ("B FM · 官方预训练权重", "BIOT (pretrained)", 3.2),
    (None, "LaBraM-Base (pretrained)", 5.8),
    (None, "CBraMod (pretrained)", 4),
    (None, "EEGPT", 4.7),
    (None, "TFM-Tokenizer", 1.9),
    ("C FM · 同 pipeline scratch", "BIOT (scratch)", 3.2),
    (None, "LaBraM-Base (scratch)", 5.8),
    (None, "CBraMod (scratch)", 4),
    (None, "PACLock (from scratch, full)", 1.64),
]

REF_NOTE = ("外部参考值(已发表,仅作校准锚,不与上表并列进论文)")
REF_BODY = ("FACED 原始论文提供数据说明与基准,但不同工作在切窗、标签聚合和跨受试者"
            "协议上差异明显。主表只放同一 pipeline 重跑结果;发表值留作校准。")

PROTOCOL = [
    ("已核查预处理协议(新集群 baseline)", None, None, None, None),
    ("冻结。2026-08-05 由 SEED-V 换回 FACED,协议取自工作簿初版。", None, None, None, None),
    ("参数", "最终值", None, None, "来源 / 执行要求"),
    ("数据版本", "FACED Synapse syn50614194;使用官方发布的 pre-processed .pkl",
     None, None, "冻结;https://doi.org/10.7303/syn50614194"),
    ("原始规模", "123 subjects(S000–S122),32 electrodes,28 videos",
     None, None, "冻结;https://www.nature.com/articles/s41597-023-02650-w"),
    ("原始采样率", "raw 为 250 或 1000 Hz;官方 pre-processed 数据统一为 250 Hz",
     None, None, "冻结;官方论文"),
    ("官方预处理", "每视频取末 30 s;统一 250 Hz;0.05–47 Hz;坏导联插值;ICA 去眼动;"
     "common-average reference;统一通道顺序", None, None,
     "冻结;官方论文。我方不得再做一次滤波"),
    ("标签", "anger, disgust, fear, sadness, neutral, amusement, inspiration, joy, "
     "tenderness 九类;视频数 3,3,3,3,4,3,3,3,3", None, None,
     "冻结;官方论文 + CBraMod label array"),
    ("官方 split", "无固定 train/val/test;原论文 cross-subject 10-fold",
     None, None, "冻结;官方论文"),
    ("benchmark split", "S000–S079 train,S080–S099 val,S100–S122 test",
     None, None, "冻结;CBraMod preprocessing_faced.py"),
    ("使用通道", "官方 pre-processed 32 通道及其发布顺序", None, None, "冻结;官方论文"),
    ("目标采样率", "250 Hz → 200 Hz", None, None, "冻结;CBraMod official code"),
    ("窗口/stride", "每个 30 s trial 切为 3 个不重叠 10 s 窗口;stride=10 s",
     None, None, "冻结;CBraMod official code"),
    ("数据泄漏约束", "先按 subject split,再切窗;同一 subject/video 的窗口只能属于一个 split",
     None, None, "冻结;代码与 benchmark 明确化"),
    ("归一化", "数值除以 100", None, None, "冻结;CBraMod loader"),
    ("类别不平衡", "不加 class weight,不重采样;unweighted CE + label smoothing 0.1",
     None, None, "冻结;CBraMod trainer"),
    ("主指标", "Balanced Accuracy;同时报告 Cohen's Kappa、Weighted F1",
     None, None, "冻结;benchmark 决策 + CBraMod evaluator"),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    out = args.out or args.xlsx.replace(".xlsx", "_filled.xlsx")
    if os.path.abspath(out) == os.path.abspath(args.xlsx):
        raise SystemExit("refusing to overwrite the input workbook")
    if not os.path.exists(out):
        shutil.copy2(args.xlsx, out)

    wb = openpyxl.load_workbook(out)
    if "FACED" in wb.sheetnames:
        print("FACED sheet 已存在,跳过")
        return
    # place it where SEED-V sat, before BCI-IV-2a, so sheet order matches the docs
    idx = wb.sheetnames.index("BCI-IV-2a") if "BCI-IV-2a" in wb.sheetnames else None
    ws = wb.create_sheet("FACED", idx)

    ws.append([TITLE]); ws["A1"].font = Font(bold=True)
    ws.append([PRIMARY])
    ws.append([])
    ws.append(HEADER)
    for c in ws[4]:
        if c.value:
            c.font = Font(bold=True)
    for grp, model, params in ROWS:
        ws.append([grp, model, params, None, None, None, None])

    ws.append([])
    ws.append([REF_NOTE]); ws.cell(row=ws.max_row, column=1).font = Font(italic=True)
    ws.append([None, REF_BODY])
    ws.append([])
    for row in PROTOCOL:
        ws.append(list(row))

    widths = {"A": 26, "B": 30, "C": 12, "D": 18, "E": 18, "F": 18, "G": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # SEED-V was dropped from the matrix on 2026-08-05 and replaced by FACED.
    # It was kept as a marked-deprecated sheet for a while so the history stayed
    # readable; it is now removed outright at the user's request. Nothing is
    # lost that matters: no SEED-V data was ever preprocessed and no run ever
    # produced a result for it, so the sheet only ever held an unused protocol.
    # The input workbook is never modified, so the original is still recoverable
    # from results/_in.xlsx.
    if "SEED-V" in wb.sheetnames:
        del wb["SEED-V"]
        print("  SEED-V sheet 已删除(该数据集不在评测矩阵内,从无数据或结果)")

    wb.save(out)
    print(f"已在 {out} 新增 FACED sheet(表头第 4 行,与其余 sheet 一致)")
    print("  A 组 5 行 / B 组 5 行 / C 组 4 行 + 外部参考值区 + 冻结协议区")


if __name__ == "__main__":
    main()
