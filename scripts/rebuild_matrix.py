"""Rebuild the results workbook to the 2026-08-24 plan (docs/DIRECTION.md).

    python -m scripts.rebuild_matrix          # -> results/PACLock_baseline_matrix.xlsx

Builds the SKELETON only -- every number is filled from runs/ by
scripts/fill_xlsx.py, so the workbook stays reproducible. Sheet order is the
paper's order of importance: paroxysmal core first, boundary corpora next,
retired corpora at the back for the record. All sheets share one layout:

  row 1 title / row 2 primary metric / row 4 header
  columns: 分组 | 模型 | 参数量 (M) | <primary> | <sec1> | <sec2> | 备注

Row labels are the contract with fill_xlsx.MODEL_ROW_LABEL: a run directory
runs/<dataset>-<variant>/ fills the row whose label matches its variant.
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

OUT = "results/PACLock_baseline_matrix.xlsx"
HDR_FILL = PatternFill("solid", fgColor="DDDDDD")
GREY = Font(color="999999", italic=True)

# sheet, title, metrics (primary first), note appended under the title
CORE = [
    ("TUEV", "TUEV — 痫样事件分类（6 类）", ["Cohen's Kappa", "Balanced Acc", "Weighted F1"], ""),
    ("TUSZ", "TUSZ — 癫痫发作检测（二分类）", ["AUC-PR", "AUROC", "Balanced Acc"], ""),
    ("CHB-MIT", "CHB-MIT — 小儿癫痫发作检测（二分类）", ["AUC-PR", "AUROC", "Balanced Acc"], ""),
    ("IIIC", "IIIC — ICU 有害脑活动分类（6 类；BDSP SPaRCNet 数据）", ["Cohen's Kappa", "Balanced Acc", "Weighted F1"],
     "数据已到（sparcnet_data npy），loader 待写；对齐 TFM-Tokenizer 的格子"),
    ("TUEP", "TUEP — 癫痫 / 非癫痫，记录级二分类", ["AUROC", "Balanced Acc", "AUC-PR"], ""),
    ("TUAB", "TUAB — 异常 EEG 检测（二分类）", ["Balanced Acc", "AUC-PR", "AUROC"], ""),
    ("ADFD", "ADFD — 阿尔茨海默 / 额颞叶痴呆 / 健康（3 类）", ["Balanced Acc", "Cohen's Kappa", "Weighted F1"], ""),
    ("CAUEEG", "CAUEEG — 痴呆三分类（韩国外部队列，dementia benchmark）", ["Balanced Acc", "Cohen's Kappa", "Weighted F1"],
     "数据已批（2026-08-24），loader 待写；领域内对照 = CEEDNet（数据集作者，NeuroImage 2023）"),
    ("Siena", "Siena — 癫痫发作检测（外部队列，PhysioNet）", ["AUC-PR", "AUROC", "Balanced Acc"],
     "数据下载中，loader 待写"),
]
BOUNDARY = [
    ("ISRUC", "ISRUC — 睡眠分期（5 类）〔边界语料：机制分析章节〕",
     ["Cohen's Kappa", "Balanced Acc", "Weighted F1"],
     "持续谱态任务：论文中作方法边界与机制分析，不进核心主张"),
    ("BCI-IV-2a", "BCI-IV-2a — 运动想象（4 类）〔边界语料：机制分析章节〕",
     ["Balanced Acc", "Cohen's Kappa", "Weighted F1"],
     "协方差域任务（功率的混合≠混合的功率）：作方法边界分析"),
]
RETIRED = [
    ("Sleep-EDF", "Sleep-EDF — 睡眠分期〔已除名〕", ["Cohen's Kappa", "Balanced Acc", "Weighted F1"],
     "除名：与 ISRUC 同族，边界分析保留 ISRUC 一个即可"),
    ("PhysioNet-MI", "PhysioNet-MI — 运动想象〔已除名〕", ["Balanced Acc", "Cohen's Kappa", "Weighted F1"],
     "除名：与 BCI-IV-2a 同族"),
    ("FACED", "FACED — 情绪（9 类）〔已除名〕", ["Balanced Acc", "Cohen's Kappa", "Weighted F1"],
     "除名：15 个变体全在随机附近（docs/FINDINGS.md）"),
    ("TUAR", "TUAR — 伪迹分类〔已除名〕", ["Cohen's Kappa", "Balanced Acc", "Weighted F1"],
     "除名：伪迹=宽带+量纲+拓扑，与振荡频带词表机制不符（FINDINGS 第五部分）"),
    ("Mumtaz2016", "Mumtaz2016 — 抑郁症检测〔已除名〕", ["AUROC", "Balanced Acc", "AUC-PR"],
     "除名：基准饱和（BIOT 0.9999），不区分模型"),
    ("EEGMat", "EEGMat — 心算/静息〔已除名〕", ["AUROC", "Balanced Acc", "AUC-PR"],
     "除名：1,199 训练窗口，小样本失效模式同 FACED"),
]

# (group label, [(variant_key, row label, static note or "")])
GROUPS = [
    ("A0 经典特征", [
        ("feat_lr",  "手工特征+LogReg", "bandpower+Hjorth+谱熵；批评文献标配"),
        ("feat_lda", "手工特征+LDA", ""),
    ]),
    ("A 轻量监督", [
        ("sparcnet", "SPaRCNet", ""),
        ("contrawr", "ContraWR", ""),
        ("cnn_transformer", "CNN-Transformer", ""),
        ("ffcl", "FFCL", ""),
        ("st_transformer", "ST-Transformer", ""),
        ("eegnet", "EEGNet (调参)", "公开搜索预算调参"),
        ("eegconformer", "EEGConformer (调参)", "benchmark 文献最强单体"),
    ]),
    ("B FM · 官方预训练权重", [
        ("biot_prest16", "BIOT (pretrained)", ""),
        ("labram_pretrained", "LaBraM-Base (pretrained)", ""),
        ("cbramod_pretrained", "CBraMod (pretrained)", ""),
        ("eegpt_pretrained", "EEGPT", ""),
        ("tfm_pretrained", "TFM-Tokenizer", "tokenizer 直接对手"),
        ("reve_pretrained", "REVE-Base (pretrained)", "新增；HF 权重"),
        ("csbrain_pretrained", "CSBrain (pretrained)", "新增；GDrive 权重"),
        ("brainomni_pretrained", "BrainOmni (pretrained)", "新增；HF 权重"),
    ]),
    ("C FM · 同 pipeline scratch", [
        ("biot_scratch", "BIOT (scratch)", ""),
        ("labram_scratch", "LaBraM-Base (scratch)", ""),
        ("cbramod_scratch", "CBraMod (scratch)", ""),
        ("eegpt_scratch", "EEGPT (scratch)", ""),
        ("tfm_scratch", "TFM-Tokenizer (scratch)", ""),
    ]),
    ("D PACLock", [
        ("paclock_duplex", "PACLock (duplex, scratch)", ""),
        ("paclock_duplex_pt2", "PACLock (duplex, 预训练 v2)", "band_norm_pac 目标"),
        ("paclock_probe_v2", "PACLock (冻结探针, v2)", "patch_len 200；表征质量协议"),
        ("paclock_probe_rand", "PACLock (冻结探针, 随机初始化)", "四道门第 4 门对照"),
    ]),
]

# ⚠ flags: (sheet, variant) -> remark overriding the static note
FLAGS = {
    ("CHB-MIT", "biot_prest16"): "⚠ 自我重叠：BIOT 预训练集含 CHB-MIT，本行不作对比依据",
    ("TUAB", "biot_prest16"):    "⚠ 自我重叠：BIOT 预训练集含 TUAB，本行不作对比依据",
}

# grey reference block: numbers self-reported by papers WITHOUT public weights
# (or different pipelines) -- context only, never comparison
REFS = {
    "TUEV": [("CodeBrain (ICLR'26, 自报)", {"Cohen's Kappa": 0.6912}),
             ("Uni-NTFM-large (ICLR'26, 自报)", {"Cohen's Kappa": 0.7030}),
             ("CBraMod (原论文自报)", {"Cohen's Kappa": 0.6512})],
    "TUAB": [("CodeBrain (ICLR'26, 自报)", {"Balanced Acc": 0.8294}),
             ("Uni-NTFM-large (ICLR'26, 自报)", {"Balanced Acc": 0.8197}),
             ("REVE-Base (原论文自报)", {"Balanced Acc": 0.8315})],
    "CAUEEG": [("CEEDNet (数据集作者, NeuroImage'23)", {})],
}


def build_sheet(ws, title, metrics, note):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "主指标：%s" % metrics[0]
    ws["A2"].font = Font(italic=True)
    if note:
        ws["A3"] = note
        ws["A3"].font = GREY
    headers = ["分组", "模型", "参数量 (M)"] + metrics + ["备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    r = 5
    for group, rows in GROUPS:
        for i, (key, label, static_note) in enumerate(rows):
            ws.cell(row=r, column=1, value=group if i == 0 else "")
            ws.cell(row=r, column=2, value=label)
            remark = FLAGS.get((ws.title, key), static_note)
            if remark:
                cell = ws.cell(row=r, column=len(headers), value=remark)
                if remark.startswith("⚠"):
                    cell.font = Font(color="CC0000")
                else:
                    cell.font = GREY
            r += 1
    refs = REFS.get(ws.title)
    if refs:
        r += 1
        ws.cell(row=r, column=1, value="E 参考值（论文自报，不同管线，不可直接比）").font = GREY
        for name, vals in refs:
            ws.cell(row=r, column=2, value=name).font = GREY
            for m, v in vals.items():
                col = 4 + metrics.index(m)
                cell = ws.cell(row=r, column=col, value="%.4f" % v)
                cell.font = GREY
            r += 1
    ws.cell(row=r + 1, column=1,
            value="空单元格 = 尚未跑；单 seed 数字标注 (1 seed)、斜体灰 —— 进论文正表需 3 seed（硬规则 4）")
    ws.cell(row=r + 1, column=1).font = GREY
    for c, wdt in enumerate([26, 30, 11] + [15] * len(metrics) + [46], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = wdt


def build_readme(ws):
    lines = [
        ("PACLock 基线矩阵（2026-08-24 重建 —— 定位:耦合感知 tokenizer，阵发性临床评测域）", True),
        ("", False),
        ("sheet 顺序 = 论文重要性：核心 9（发作/痫样/临床判别/痴呆）→ 边界 2（机制分析）→ 已除名 6（存档）", False),
        ("核心：TUEV TUSZ CHB-MIT IIIC TUEP TUAB ADFD CAUEEG Siena", False),
        ("边界：ISRUC BCI-IV-2a（输的机制有文献支撑，作『方法边界』章节，不进核心主张）", False),
        ("除名：Sleep-EDF PhysioNet-MI FACED TUAR Mumtaz2016 EEGMat（理由见各 sheet 标题与 docs/FINDINGS.md）", False),
        ("", False),
        ("行分组", True),
        ("A0 经典特征基线 —— 2026 批评文献的标配（手工特征在多个临床任务打平/打赢 FM）", False),
        ("A  轻量监督（BIOT 配方五件套 + 调参 EEGNet/Conformer；调参需公开搜索预算）", False),
        ("B  FM · 官方预训练权重（复现门：先在重叠语料复现其发表数字）", False),
        ("C  同 pipeline from-scratch（与 B 同架构，剥离预训练贡献）", False),
        ("缓议：MOMENT 等通用时序 FM 对照 —— 不承重于 tokenizer 主张，留作 rebuttal 期选项", False),
        ("D  PACLock（scratch / 预训练 v2 / 冻结探针 vs 随机初始化探针 —— 四道门第 4 门）", False),
        ("E  参考值（灰）：无公开权重或异管线论文自报数，只作上下文", False),
        ("", False),
        ("硬规则", True),
        ("1. B 组任一模型进表前须先复现其发表数字（噪声内）", False),
        ("2. 每个模型用它自己发表的配方；不为任何 baseline 发明超参", False),
        ("3. 配置错误的 seed 拒绝写入", False),
        ("4. 论文正表 = 3 seed 均值±标准差；单 seed 只以 (1 seed) 斜体灰入表", False),
        ("5. BIOT 在 CHB-MIT/TUAB 的行带 ⚠ 自我重叠旗标，不作对比依据", False),
        ("", False),
        ("填表：scripts/rebuild_matrix.py 造骨架 → scripts/fill_xlsx.py 从 runs/ 回填（行标签=运行目录 variant）", False),
        ("负对照套件（身份探针/LEACE/标签置换/配对 CI）另见 _消融 与论文附录，不入本表", False),
    ]
    for i, (t, bold) in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=t).font = Font(bold=bold)
    ws.column_dimensions["A"].width = 110


def main():
    wb = openpyxl.Workbook()
    build_readme(wb.active)
    wb.active.title = "README"
    for sheet, title, metrics, note in CORE + BOUNDARY + RETIRED:
        build_sheet(wb.create_sheet(sheet), title, metrics, note)
    wb.save(OUT)
    print("rebuilt", OUT, "with", len(wb.sheetnames), "sheets:")
    print(" ", wb.sheetnames)


if __name__ == "__main__":
    main()
