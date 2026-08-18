"""Add the two 'PACLock (pretrained, ...)' rows to every sheet of the matrix.

Inserted directly under 'PACLock (from scratch, full)' so each pretrained
result reads against its own from-scratch baseline -- the pair differs only
by the checkpoint, which is the whole comparison.

Formatting (font/border/fill/alignment/number format) is copied cell-by-cell
from the from-scratch PACLock row rather than left default: normalize_xlsx.py
redraws the results block on a uniform grid and a row that arrives with no
styling reads as "not part of the table" to it.

Idempotent -- re-running finds the labels already present and does nothing.
"""
import copy
import sys

import openpyxl

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"
ANCHOR = "PACLock (from scratch, full)"
NEW_ROWS = ["PACLock (pretrained, base)", "PACLock (pretrained, large)"]

wb = openpyxl.load_workbook(XLSX)
touched = []
for name in wb.sheetnames:
    if name in ("README", "_填写记录"):
        continue
    ws = wb[name]

    anchor_row = model_col = None
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == ANCHOR:
                anchor_row, model_col = c.row, c.column
                break
        if anchor_row:
            break
    if anchor_row is None:
        print("  %-14s no PACLock anchor row -- skipped" % name)
        continue

    existing = {ws.cell(row=r, column=model_col).value
                for r in range(1, ws.max_row + 1)}
    todo = [lbl for lbl in NEW_ROWS if lbl not in existing]
    if not todo:
        print("  %-14s already has the pretrained rows" % name)
        continue

    ws.insert_rows(anchor_row + 1, amount=len(todo))
    for i, label in enumerate(todo):
        r = anchor_row + 1 + i
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=anchor_row, column=col)
            dst = ws.cell(row=r, column=col)
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.border = copy.copy(src.border)
                dst.fill = copy.copy(src.fill)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format
        ws.cell(row=r, column=model_col, value=label)
    touched.append((name, anchor_row, len(todo)))
    print("  %-14s inserted %d row(s) after row %d" % (name, len(todo), anchor_row))

if touched:
    wb.save(XLSX)
    print("saved %s" % XLSX)
else:
    print("nothing to do")
