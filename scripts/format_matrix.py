"""Apply the workbook's reading conventions after a fill.

    python -m scripts.format_matrix [--xlsx results/PACLock_baseline_matrix_filled.xlsx]

Four rules, applied per dataset sheet:

1. The 备注 column is dropped -- the conventions live in README, not per row.
2. GREY font marks a cell whose recipe we transplanted: the baseline's own
   paper never ran that corpus, so its hyperparameters are ours, not the
   authors'. A reader must be able to see at a glance which numbers rest on
   a published recipe and which rest on our transplant (hard rule 2).
   PUBLISHED_ON below is the per-model list of OUR nine corpora that the
   model's own paper evaluated.
3. YELLOW fill marks a cell with fewer than three seeds -- exploration-grade,
   not paper-table-grade (hard rule 4). Our own rows are single-seed by
   policy, so they are yellow by construction.
4. BOLD marks the best value in each metric column. The E-block (self-
   reported numbers from other papers, different pipelines) is excluded from
   that comparison -- bolding across pipelines would assert a comparison the
   protocol forbids.
"""
from __future__ import annotations

import argparse
import re

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_TO_DS = {
    "TUEV": "tuev", "TUSZ": "tusz", "CHB-MIT": "chbmit", "IIIC": "iiic",
    "TUEP": "tuep", "TUAB": "tuab", "ADFD": "adfd", "CAUEEG": "caueeg",
    "Siena": "siena", "ISRUC": "isruc", "BCI-IV-2a": "bci_iv_2a",
    "Sleep-EDF": "sleepedf", "PhysioNet-MI": "physionet_mi", "FACED": "faced",
    "TUAR": "tuar", "Mumtaz2016": "mumtaz", "EEGMat": "eegmat",
}

# Which of OUR corpora each baseline's own paper evaluated. Anything outside
# this set is a transplanted recipe -> grey.
#   group A (SPaRCNet/ContraWR/CNN-T/FFCL/ST-T) come from BIOT's own training
#   scripts, so their published coverage is BIOT's evaluation set.
BIOT_SET = {"tuab", "tuev", "chbmit", "iiic"}
PUBLISHED_ON = {
    "手工特征 (LR/LDA 取优)": set(),          # ours; no author recipe exists
    "SPaRCNet": BIOT_SET, "ContraWR": BIOT_SET, "CNN-Transformer": BIOT_SET,
    "FFCL": BIOT_SET, "ST-Transformer": BIOT_SET,
    "EEGNet (调参)": set(),                   # tuned by us on a declared grid
    "EEGConformer (调参)": set(),
    "BIOT (pretrained)": BIOT_SET, "BIOT (scratch)": BIOT_SET,
    "LaBraM-Base (pretrained)": {"tuab", "tuev"},
    "LaBraM-Base (scratch)": {"tuab", "tuev"},
    "CBraMod (pretrained)": {"tuab", "tuev", "chbmit"},
    "CBraMod (scratch)": {"tuab", "tuev", "chbmit"},
    "EEGPT": {"tuab", "tuev"}, "EEGPT (scratch)": {"tuab", "tuev"},
    "TFM-Tokenizer": {"tuab", "tuev", "chbmit", "iiic"},
    "TFM-Tokenizer (scratch)": {"tuab", "tuev", "chbmit", "iiic"},
    "REVE-Base (pretrained)": {"tuab", "tuev"},
    "CSBrain (pretrained)": {"tuab", "tuev", "chbmit", "siena"},
}
OURS_PREFIX = "PACLock"

GREY = Font(color="8C8C8C")
BLACK = Font(color="000000")
GREY_BOLD = Font(color="8C8C8C", bold=True)
BLACK_BOLD = Font(bold=True)
YELLOW = PatternFill("solid", fgColor="FFF2CC")
NUM = re.compile(r"^\s*(-?\d*\.?\d+)")


def value_of(cell) -> float | None:
    if cell.value is None:
        return None
    m = NUM.match(str(cell.value))
    return float(m.group(1)) if m else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="results/PACLock_baseline_matrix_filled.xlsx")
    args = ap.parse_args()
    wb = openpyxl.load_workbook(args.xlsx)

    for sheet, ds in SHEET_TO_DS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        # --- 1. drop the 备注 column -------------------------------------
        hdr = 4
        remark_col = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=hdr, column=c).value or "").strip() == "备注":
                remark_col = c
                break
        if remark_col:
            ws.delete_cols(remark_col)

        metric_cols = [c for c in range(4, ws.max_column + 1)
                       if ws.cell(row=hdr, column=c).value]
        best = {c: (None, None) for c in metric_cols}   # col -> (value, row)
        in_reference = False

        for r in range(hdr + 1, ws.max_row + 1):
            group = str(ws.cell(row=r, column=1).value or "")
            label = str(ws.cell(row=r, column=2).value or "").strip()
            if group.startswith("E 参考值"):
                in_reference = True
            if not label:
                continue

            ours = label.startswith(OURS_PREFIX)
            # --- 2. grey where the recipe is transplanted ----------------
            transplanted = (not ours and not in_reference
                            and ds not in PUBLISHED_ON.get(label, set()))

            for c in metric_cols:
                cell = ws.cell(row=r, column=c)
                if cell.value in (None, ""):
                    continue
                cell.font = GREY if (transplanted or in_reference) else BLACK
                # --- 3. yellow for <3 seeds -----------------------------
                if "seed" in str(cell.value):
                    cell.fill = YELLOW
                # --- 4. track the column best (E-block excluded) --------
                if not in_reference:
                    v = value_of(cell)
                    if v is not None and (best[c][0] is None or v > best[c][0]):
                        best[c] = (v, r)

        for c, (v, r) in best.items():
            if r is None:
                continue
            cell = ws.cell(row=r, column=c)
            cell.font = GREY_BOLD if cell.font.color and \
                cell.font.color.rgb and "8C8C8C" in str(cell.font.color.rgb) \
                else BLACK_BOLD
        widths = [26, 30, 11] + [16] * len(metric_cols)
        for c, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = wdt
        print("  %-12s %d metric columns formatted" % (sheet, len(metric_cols)))

    # --- README: state the conventions now that 备注 is gone -------------
    ws = wb["README"]
    row = ws.max_row + 2
    for text, bold in (
        ("单元格读法（2026-08-26 起）", True),
        ("灰色数字 = 该 baseline 的原论文没有跑过这个语料，超参是我们移植的（硬规则 2 的可见化）", False),
        ("黄色底 = 不足 3 seed，探索级而非进表级（硬规则 4）；我们自己的行按 seed 纪律恒为单 seed", False),
        ("加粗 = 该列最好成绩；E 组参考值（他人论文自报、异管线）不参与加粗比较", False),
        ("空格 = 尚未跑或有意不排（见 README 上方缓议清单）", False),
    ):
        ws.cell(row=row, column=1, value=text).font = Font(bold=bold)
        row += 1

    wb.save(args.xlsx)
    print("saved", args.xlsx)


if __name__ == "__main__":
    main()
