"""Fill three delta columns per sheet: vs PACLock from-scratch, vs pretrained
base, vs pretrained large.

Replaces the single "Δ<metric> vs PACLock" column. One reference is no longer
enough now that PACLock has three rows -- a reader comparing a baseline to
"PACLock" needs to know which PACLock, and the base-vs-large question is
answered by reading the same row twice.

Rewrites the header for the first delta column too (it said "vs PACLock",
ambiguous once there are three), so this is safe to re-run: it locates the
delta block by position after the metric columns rather than by matching the
old header text.

Deltas are (row's value - reference value) in the sheet's own primary metric,
so a positive number always means "this row beats that PACLock variant".
Reference rows that are themselves empty leave their whole column blank
rather than diffing against a stale number.
"""
import re

import openpyxl

from scripts.fill_xlsx import METRIC_HEADER, MODEL_ROW_LABEL, find_results_block

XLSX = "results/PACLock_baseline_matrix_filled.xlsx"
VALUE_RE = re.compile(r"^(-?\d+\.\d+)")

REFS = [
    ("paclock_v2", "vs scratch"),
    ("paclock_pt_base", "vs pt-base"),
    ("paclock_pt_large", "vs pt-large"),
]


def parse_mean(v):
    if not isinstance(v, str):
        return None
    m = VALUE_RE.match(v.strip())
    return float(m.group(1)) if m else None


wb = openpyxl.load_workbook(XLSX)
total = 0
for name in wb.sheetnames:
    if name in ("README", "_填写记录"):
        continue
    ws = wb[name]
    hdr_row, model_col, cols, _ = find_results_block(ws)
    if hdr_row is None:
        print("  %-14s no results block" % name)
        continue

    # the delta block starts right after the last metric column
    first_delta = max(cols.values()) + 1
    # the sheet's primary metric is the one its existing delta column named;
    # fall back to the last metric column if that header is gone
    existing = ws.cell(row=hdr_row, column=first_delta).value
    metric = None
    if isinstance(existing, str) and existing.strip().startswith("Δ"):
        cand = existing.lstrip("Δ").split(" vs ")[0].strip()
        if cand in cols:
            metric = cand
    if metric is None:
        metric = max(cols, key=lambda k: cols[k])
    metric_col = cols[metric]

    # locate each reference row
    label_to_row = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=model_col).value
        if isinstance(v, str) and v.strip():
            label_to_row[v.strip()] = r

    for i, (key, short) in enumerate(REFS):
        col = first_delta + i
        ws.cell(row=hdr_row, column=col, value="Δ%s %s" % (metric, short))
        ref_label = MODEL_ROW_LABEL[key]
        ref_row = label_to_row.get(ref_label)
        ref_val = parse_mean(ws.cell(row=ref_row, column=metric_col).value) if ref_row else None
        if ref_val is None:
            print("  %-14s %-10s reference %r empty -- column blank"
                  % (name, short, ref_label))
            for r in range(hdr_row + 1, ws.max_row + 1):
                ws.cell(row=r, column=col, value=None)
            continue
        n = 0
        for r in range(hdr_row + 1, ws.max_row + 1):
            lbl = ws.cell(row=r, column=model_col).value
            if not isinstance(lbl, str) or not lbl.strip():
                continue
            if r == ref_row:
                ws.cell(row=r, column=col, value="—")
                continue
            v = parse_mean(ws.cell(row=r, column=metric_col).value)
            ws.cell(row=r, column=col, value=("%+.4f" % (v - ref_val)) if v is not None else None)
            if v is not None:
                n += 1
        total += n
        print("  %-14s %-10s ref=%.4f -> %d deltas" % (name, short, ref_val, n))

wb.save(XLSX)
print("\n%d delta cells written across 3 columns per sheet" % total)
